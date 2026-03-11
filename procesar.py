import pdfplumber
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def limpiar_monto(texto):
    """Extrae solo los números y la coma decimal de una cadena de texto."""
    if not texto:
        return 0.0
    # Busca el primer grupo de números que puede tener puntos de miles y coma decimal
    # Ejemplo: "$ 2.500.000,00" -> "2500000,00"
    encontrado = re.search(r'([\d\.]+,\d{2})|(\d+)', texto)
    if encontrado:
        valor = encontrado.group().replace(".", "").replace(",", ".")
        try:
            return float(valor)
        except:
            return 0.0
    return 0.0

def ejecutar_reto_unificado():
    ruta_entrada = "PDFs/"
    base_datos = []

    if not os.path.exists(ruta_entrada):
        print(f"⚠️ Error: No encuentro la carpeta '{ruta_entrada}'.")
        return

    print("--- Iniciando extracción de datos (Versión Ultra-Robusta) ---")

    for archivo in os.listdir(ruta_entrada):
        if archivo.endswith(".pdf"):
            try:
                with pdfplumber.open(os.path.join(ruta_entrada, archivo)) as pdf:
                    texto_completo = ""
                    for pagina in pdf.pages:
                        texto_completo += pagina.extract_text() + "\n"
                    
                    lineas = texto_completo.split('\n')
                    
                    datos = {
                        "Compromiso": "", "Nombre": "", "ID": "",
                        "Bruto": 0.0, "ICA": 0.0, "USO": "A-02-02-02-009-002-09",
                        "Mes a pagar": "Febrero"
                    }

                    for linea in lineas:
                        # 1. Compromiso SIIF
                        if "Compromiso SIIF" in linea:
                            match = re.search(r'SIIF\s+(\d+)', linea)
                            if match: datos["Compromiso"] = match.group(1)

                        # 2. Nombre
                        if "Nombres y apellidos:" in linea:
                            datos["Nombre"] = linea.split("apellidos:")[1].split("Banco")[0].strip()
                        
                        # 3. Cédula
                        if "Cédula de Ciudadanía" in linea:
                            match = re.search(r'Ciudadanía\s+([\d\.]+)', linea)
                            if match: datos["ID"] = match.group(1).replace(".", "")
                        
                        # 4. Valor Bruto Pago
                        if "Valor Bruto Pago:" in linea:
                            monto_texto = linea.split("Pago:")[1].split("Saldo")[0]
                            datos["Bruto"] = limpiar_monto(monto_texto)
                        
                        # 5. Reteica - 8299
                        if "Reteica - 8299" in linea:
                            monto_texto = linea.split("MANIZALES")[1]
                            datos["ICA"] = limpiar_monto(monto_texto)

                    datos["Total"] = datos["Bruto"] - datos["ICA"]
                    
                    if datos["Nombre"]: # Solo agregar si encontró al menos el nombre
                        base_datos.append(datos)
                        print(f"✅ Extraído: {datos['Nombre']} | Compromiso: {datos['Compromiso']} | Bruto: {datos['Bruto']}")
                    
            except Exception as e:
                print(f"❌ Error en {archivo}: {e}")

    if base_datos:
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidado AVAL"

        # Encabezados Superiores
        ws.merge_cells('A1:B1'); ws['A1'] = "Nombre supervisor"
        ws.merge_cells('J1:K1'); ws['J1'] = "Número de identificación"
        ws.merge_cells('A2:B2'); ws['A2'] = "Dependencia"
        ws.merge_cells('C2:I2'); ws['C2'] = "Centro de Procesos Industriales y de la Construcción"
        ws.merge_cells('J2:K2'); ws['J2'] = "Fecha de la solicitud"
        ws.merge_cells('L2:N2'); ws['L2'] = "FEBRERO"

        headers = ["No. Consecutivo", "No. Compromiso ptal.", "USO", "Mes a pagar", 
                   "Nombre del contratista", "Tipo de identificación", "Número de Identificación", 
                   "Valor bruto de la obligación", "Valor Retención ICA", "Valor Retención Fuente", 
                   "Valor Retención IVA", "Valor Retención Embargos", "Valor otras Retenciones", "Total a pagar"]
        ws.append(headers)

        for cell in ws[3]:
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Llenar datos
        for i, d in enumerate(base_datos, start=1):
            fila = [i, d["Compromiso"], d["USO"], d["Mes a pagar"], d["Nombre"], "CC", d["ID"],
                    d["Bruto"], d["ICA"], 0, 0, 0, 0, d["Total"]]
            ws.append(fila)
            
            idx = ws.max_row
            # Formato de miles y colores
            for col in range(8, 15):
                ws.cell(row=idx, column=col).number_format = '#,##0'
            
            ws[f'H{idx}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws[f'I{idx}'].fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
            ws[f'N{idx}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        # Ajustar ancho
        for l in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[l].width = 18

        wb.save("RESULTADO_AVAL_FEBRERO.xlsx")
        print("\n--- 🚀 PROCESO COMPLETADO ---")
    else:
        print("⚠️ No se pudo extraer información válida de los PDFs.")

if __name__ == "__main__":
    ejecutar_reto_unificado()