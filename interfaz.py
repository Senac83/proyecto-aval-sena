# Desarrollado por: Mariana Ramirez Giraldo
# Ficha ADSO: 3063934

import streamlit as st
import pdfplumber
import re
import io
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image

# --- CONFIGURACIÓN DE PÁGINA Y FAVICON ---
ruta_logo = "img/lgo_sena.png"

if os.path.exists(ruta_logo):
    SENA_LOGO = Image.open(ruta_logo)
else:
    SENA_LOGO = None

st.set_page_config(
    page_title="SENA | AVAL", 
    page_icon=SENA_LOGO if SENA_LOGO else "📄", 
    layout="centered",
    initial_sidebar_state="expanded" 
)

# --- CSS INTEGRAL ---
st.markdown(f"""
<style>
    .stApp {{ background-color: #FFFFFF !important; }}

    [data-testid="stWidgetLabel"] p {{
        color: #000000 !important;
        font-weight: bold !important;
        font-size: 16px !important;
    }}

    [data-testid="stFileUploader"] section {{
        background-color: #F0F9EB !important;
        border: 2px dashed #39A900 !important;
        border-radius: 15px !important;
    }}

    [data-testid="stFileUploader"] section div div span,
    [data-testid="stFileUploader"] section div div small {{
        color: #39A900 !important;
        font-weight: bold !important;
    }}

    [data-testid="stFileUploaderFileName"] {{
        color: #1b5e20 !important; 
        font-weight: bold !important;
    }}

    [data-testid="stFileUploaderFileData"] small {{
        color: #39A900 !important;
        opacity: 1 !important;
    }}

    [data-testid="stFileUploader"] button {{
        background-color: #39A900 !important;
        color: white !important;
    }}

    .stButton>button, .stDownloadButton>button {{
        background-color: #39A900 !important;
        color: #FFFFFF !important;
        border-radius: 30px !important;
        border: none !important;
        font-weight: bold !important;
    }}

    [data-testid="stSidebar"] {{
        background-color: #39A900 !important;
    }}
    
    [data-testid="stSidebar"] * {{
        color: #FFFFFF !important;
    }}

    [data-testid="stSidebarCollapsedControl"] svg {{
        fill: #39A900 !important;
        color: #39A900 !important;
    }}
    
    [data-testid="stSidebar"] [role="button"] svg {{
        fill: #39A900  !important;
        color: #39A900  !important;
    }}

    

    [data-testid="stFileUploader"] button[kind="icon"] {{
        transform: scale(1.2);
    }}

    /*  BARRA DE PROGRESO  */
    [data-testid="stProgress"] > div > div > div > div {{
        background-color: #39A900 !important;
    }}

    
    [data-testid="stProgress"] {{
        height: 10px !important;
    }}

    header[data-testid="stHeader"] {{ background: transparent !important; }}
    .stDeployButton {{ display:none !important; }}
</style>
""", unsafe_allow_html=True)

# --- LÓGICA DE ESTADO ---
if 'uploader_key' not in st.session_state: st.session_state.uploader_key = 0
if 'datos_listos' not in st.session_state: st.session_state.datos_listos = None

def reiniciar_todo():
    st.session_state.uploader_key += 1
    st.session_state.datos_listos = None

def limpiar_monto(texto):
    if not texto: return 0.0
    encontrado = re.search(r'([\d\.]+,\d{2})|(\d+)', texto)
    if encontrado:
        valor = encontrado.group().replace(".", "").replace(",", ".")
        try: return float(valor)
        except: return 0.0
    return 0.0

# --- BARRA LATERAL ---
with st.sidebar:
    if SENA_LOGO:
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.image(SENA_LOGO, use_container_width=True)
            
    st.markdown("### Información")
    st.write(f"**Aprendiz:** Mariana Ramirez G.")
    st.write(f"**Ficha:** 3063934")
    st.write(f"**Instructor:** Carlos Andres Loaiza R.")
    st.write("---")
    st.write("Centro de Procesos Industriales")

# --- CUERPO PRINCIPAL ---
st.markdown("<h1 style='color: #39A900; text-align: center; margin-bottom: 0;'>SENA | AVAL</h1>", unsafe_allow_html=True)
st.markdown("<p style='text-align: center; color: #000; font-weight: 500;'>Procesador de extractos PDF</p>", unsafe_allow_html=True)
st.write("---")

uploaded_files = st.file_uploader(
    "Selecciona los archivos PDF", 
    type="pdf", 
    accept_multiple_files=True,
    key=f"uploader_{st.session_state.uploader_key}"
)

# Botón de Procesar
if uploaded_files and st.session_state.datos_listos is None:
    st.write("##")
    if st.button("⚙️ PROCESAR DOCUMENTOS"):
        base_datos = []

        progress_bar = st.progress(0)
        total_archivos = len(uploaded_files)

        with st.spinner("Analizando información..."):
            for i, uploaded_file in enumerate(uploaded_files, start=1):
                with pdfplumber.open(uploaded_file) as pdf:
                    texto = "\n".join([p.extract_text() for p in pdf.pages if p.extract_text()])
                    lineas = texto.split("\n")
                    datos = {"USO": "A-02-02-02-009-002-09", "Mes": "MES", "Bruto": 0.0, "ICA": 0.0}
                    for linea in lineas:
                        if "Fecha Elaboración" in linea:
                            m_fecha = re.search(r'Fecha Elaboración\s+([a-zA-ZáéíóúÁÉÍÓÚ]+\s+de\s+\d{4})', linea)
                            if m_fecha: datos["Mes"] = m_fecha.group(1).split(" de ")[0].capitalize()
                        if "Compromiso SIIF" in linea:
                            m = re.search(r"SIIF\s+(\d+)", linea)
                            if m: datos["Compromiso"] = m.group(1)
                        if "Nombres y apellidos:" in linea:
                            datos["Nombre"] = linea.split("apellidos:")[1].split("Banco")[0].strip()
                        if "Cédula de Ciudadanía" in linea:
                            m = re.search(r"Ciudadanía\s+([\d\.]+)", linea)
                            if m: datos["ID"] = m.group(1).replace(".", "")
                        if "Valor Bruto Pago:" in linea:
                            datos["Bruto"] = limpiar_monto(linea.split("Pago:")[1].split("Saldo")[0])
                        if "Reteica - 8299" in linea:
                            partes = linea.split("MANIZALES")
                            if len(partes) > 1: datos["ICA"] = limpiar_monto(partes[1])
                    if datos.get("Nombre"): base_datos.append(datos)

                progress_bar.progress(i / total_archivos)

        st.session_state.datos_listos = base_datos
        st.rerun()

# Botón de Descarga
if st.session_state.datos_listos:
    st.write("---")
    st.success(f"✅ ¡Listo! {len(st.session_state.datos_listos)} archivos procesados.")
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["No. Consecutivo", "No. Compromiso ptal.", "USO", "Mes a pagar", "Nombre del contratista", "ID", "Número de Identificación", "Valor bruto", "Valor ICA", "Ret_F", "Ret_IVA", "Ret_Emb", "Otras", "Total"])
    
    for i, d in enumerate(st.session_state.datos_listos, start=1):
        idx = ws.max_row + 1
        ws.append([i, d.get("Compromiso"), d.get("USO"), d.get("Mes"), d.get("Nombre"), "CC", d.get("ID"), d.get("Bruto"), d.get("ICA"), 0, 0, 0, 0, f"=H{idx}-I{idx}"])
        
        ws[f'H{idx}'].fill = PatternFill(start_color="FFFF00", fill_type="solid")
        ws[f'I{idx}'].fill = PatternFill(start_color="00B0F0", fill_type="solid")
        ws[f'N{idx}'].fill = PatternFill(start_color="92D050", fill_type="solid")

        ws[f'H{idx}'].number_format = '#,##0.00'
        ws[f'I{idx}'].number_format = '#,##0.00'
        ws[f'N{idx}'].number_format = '#,##0.00'
    
    wb.save(output)
    st.download_button(
        label="📥 DESCARGAR REPORTE EXCEL", 
        data=output.getvalue(), 
        file_name="REPORTE_AVAL_SENA.xlsx", 
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        on_click=reiniciar_todo
    )
